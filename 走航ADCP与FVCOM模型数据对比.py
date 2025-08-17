# 2025.07.17 最新，以此为准

from datetime import datetime,timedelta
import numpy as np
import openpyxl
import ast
import netCDF4
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
from geopy.distance import geodesic as gd
import os
import shutil
import matplotlib.colors as mcolors
from matplotlib.ticker import MultipleLocator
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 判断输出目录是否存在
if os.path.exists('output_compare_adcp'):
    shutil.rmtree('output_compare_adcp')
os.makedirs('output_compare_adcp')

# 读取标准文件数据
with open(r'D:\Documents\code\Pycharm\FVCOM_analysis\ADCP.txt','r') as fin:
    i=0
    fin.seek(0, 2)  # 将指针移动到文件末尾
    eof_position = fin.tell()  # 获取EOF位置
    fin.seek(0)  # 回到文件开头
    obs_datetime = []
    obs_lat = []
    obs_lon = []
    obs_tmp = []
    rcd_dep = []
    rcd_spd = []
    rcd_dir = []
    rcd_ver_spd = []
    rcd_ver_dif = []
    while fin.tell() < eof_position:
        try:
            record_lines = int(fin.read(3)) # 层数
        except ValueError:
            break
        obs_type = fin.read(2).strip() # 观测类型 0走航
        pcs_type = fin.read(2).strip() # 资料处理类型 0底跟踪
        obs_date = fin.read(9).strip() # 日期
        obs_time = fin.read(7).strip() # 时间
        obs_datetime.append(datetime.strptime(obs_date+obs_time, "%Y%m%d%H%M%S"))
        obs_lat_str = fin.read(9).strip().split(' ') # 纬度
        obs_lat.append(int(obs_lat_str[0][:2]) + float(obs_lat_str[0][2:4])/60. + float(obs_lat_str[0][4:6])/3600.)
        obs_lon_str = fin.read(10).strip().split(' ') # 经度
        obs_lon.append(int(obs_lon_str[0][:3]) + float(obs_lon_str[0][3:5])/60. + float(obs_lon_str[0][5:7])/3600.)
        obs_spd = float(fin.read(7).strip()) # 船速
        obs_dir = float(fin.read(7).strip()) # 航向
        obs_dep = float(fin.read(8).strip()) # 水深
        obs_tmp.append(float(fin.read(7).strip())) # 水温
        rcd_dep.append([])
        rcd_spd.append([])
        rcd_dir.append([])
        rcd_ver_spd.append([])
        rcd_ver_dif.append([])
        for line in range(record_lines):
            rcd_dep[i].append(float(fin.read(7))) # 观测深度
            rcd_type = fin.read(2) # 0实测 1插值
            rcd_spd[i].append(float(fin.read(6))) # 实际流速
            rcd_dir[i].append(float(fin.read(7))) # 实际流向
            rcd_ver_spd[i].append(float(fin.read(6))) # 垂直流速 向上为正向下为负
            rcd_ver_dif[i].append(float(fin.read(6))) # 垂直流速误差
            rcd_ok = int(fin.read(4)) # 良好比例
            rcd_cgl = fin.read(1) # 换行
        i+=1
# 计算深度unique
rcd_dep_all = []
for i in range(len(rcd_dep)):
    for j in range(len(rcd_dep[i])):
        rcd_dep_all.append(rcd_dep[i][j])
rcd_dep_uni = np.unique(np.array(rcd_dep_all))
# 输出到excel
workbook_out = openpyxl.Workbook()
worksheet_out = workbook_out.active
worksheet_out.cell(row=1, column=1).value = '时间'
worksheet_out.cell(row=2, column=1).value = '纬度'
worksheet_out.cell(row=3, column=1).value = '经度'
worksheet_out.cell(row=4, column=1).value = '深度(m)'
for i_row in range(len(rcd_dep_uni)):
    worksheet_out.cell(row=5+i_row, column=1).value = rcd_dep_uni[i_row]
for i_column in range(len(obs_lon)):
    worksheet_out.cell(row=1, column=2+i_column).value = obs_datetime[i_column]
    worksheet_out.cell(row=2, column=2+i_column).value = obs_lat[i_column]
    worksheet_out.cell(row=3, column=2+i_column).value = obs_lon[i_column]
    i = 0
    for i_row in range(5,worksheet_out.max_row+1):
        if worksheet_out.cell(row=i_row, column=1).value == rcd_dep[i_column][i]:
            worksheet_out.cell(row=i_row, column=2 + i_column).value = str((round(rcd_spd[i_column][i],3),round(rcd_dir[i_column][i],3)))
            if i==len(rcd_spd[i_column])-1:
                break
            else:
                i+=1
worksheet_out.title = 'horizontal speed'
# 创建垂直速度表
worksheet_out = workbook_out.create_sheet()
worksheet_out.title = 'vertical speed'
worksheet_out.cell(row=1, column=1).value = '时间'
worksheet_out.cell(row=2, column=1).value = '纬度'
worksheet_out.cell(row=3, column=1).value = '经度'
worksheet_out.cell(row=4, column=1).value = '深度(m)'
for i_row in range(len(rcd_dep_uni)):
    worksheet_out.cell(row=5+i_row, column=1).value = rcd_dep_uni[i_row]
for i_column in range(len(obs_lon)):
    worksheet_out.cell(row=1, column=2+i_column).value = obs_datetime[i_column]
    worksheet_out.cell(row=2, column=2+i_column).value = obs_lat[i_column]
    worksheet_out.cell(row=3, column=2+i_column).value = obs_lon[i_column]
    i = 0
    for i_row in range(5,worksheet_out.max_row+1):
        if worksheet_out.cell(row=i_row, column=1).value == rcd_dep[i_column][i]:
            worksheet_out.cell(row=i_row, column=2 + i_column).value = rcd_ver_spd[i_column][i]
            if i==len(rcd_spd[i_column])-1:
                break
            else:
                i+=1
workbook_out.save('obs_ADCP.xlsx')
workbook_out.close()

# 计算速度
workbook = openpyxl.load_workbook('obs_ADCP.xlsx')
sheet = workbook.active
lat = []
lon = []
for cell in sheet[2]: # 纬度
    lat.append(cell.value)
for cell in sheet[3]: # 经度
    lon.append(cell.value)
lat = lat[1:]
lon = lon[1:]
depth_adcp = []
speed_data = []
i=0
i_row = 5
while i_row < sheet.max_row:
    if abs(sheet.cell(row=i_row+1, column=1).value-sheet.cell(row=i_row, column=1).value)<0.11: # ==0.1
        depth_adcp.append(sheet.cell(row=i_row, column=1).value)
        speed_data.append([])
        for i_column in range(2,sheet.max_column+1):
            if sheet.cell(row=i_row, column=i_column).value is not None: # 缺测
                speed_data[i].append(sheet.cell(row=i_row, column=i_column).value)
            else:
                if sheet.cell(row=i_row + 1, column=i_column).value is not None: # 缺测
                    speed_data[i].append(sheet.cell(row=i_row + 1, column=i_column).value)
                else:
                    speed_data[i].append(np.nan)
        i_row += 2
        i += 1
    else:
        depth_adcp.append(sheet.cell(row=i_row, column=1).value)
        speed_data.append([])
        for i_column in range(2,sheet.max_column+1):
            value = sheet.cell(row=i_row, column=i_column).value
            if value is None:
                speed_data[i].append(np.nan)
            else:
                speed_data[i].append(sheet.cell(row=i_row, column=i_column).value)
        i_row+=1
        i+=1
u = np.zeros_like(speed_data,dtype=float)
v = np.zeros_like(speed_data,dtype=float)
u[:,:] = np.nan
v[:,:] = np.nan
for i in range(len(speed_data)):
    for j in range(len(speed_data[0])):
        try:
            np.isnan(speed_data[i][j])
        except TypeError:
            speed = ast.literal_eval(speed_data[i][j])[0]
            direction = ast.literal_eval(speed_data[i][j])[1]
            theta_rad = np.deg2rad(direction)
            u[i,j] = speed*np.sin(theta_rad)
            v[i,j] = speed*np.cos(theta_rad)
        else:
            u[i, j] = np.nan
            v[i, j] = np.nan

# 筛选经纬度 经度(129,132)
i_start = np.nan
i_end = np.nan
bool_start = False
for i_lon in range(len(lon)): # 筛选
    if bool_start==False and lon[i_lon]>129.:
        i_start = i_lon
        bool_start = True
    if bool_start and lon[i_lon]>132.:
        i_end = i_lon
        break

# 后续应当用到的数据(67层,151个点)
lon_adcp = lon[i_start:i_end]
lat_adcp = lat[i_start:i_end]
# depth_adcp = depth_adcp
u_adcp = u[:,i_start:i_end]
v_adcp = v[:,i_start:i_end]

# 读取nc数据
dataset = netCDF4.Dataset(r'D:\Documents\code\Pycharm\FVCOM_analysis\kuroshio_2024_0717\kuroshio_2024_0717.nc')
nv   = dataset.variables['nv'][:]
time = dataset.variables['time'][:]
lat  = dataset.variables['lat'][:]
lon  = dataset.variables['lon'][:]
latc = dataset.variables['latc'][:]
lonc = dataset.variables['lonc'][:]
u = dataset.variables['u'][:]
v = dataset.variables['v'][:]
zeta = dataset.variables['zeta'][:]
h = dataset.variables['h'][:]
depth_all = h+zeta # (时间，格点)
siglay = dataset.variables['siglay'][:]
depth = np.zeros((zeta.shape[0],siglay.shape[0],siglay.shape[1]),dtype=float) #(时间，层，格点)
for i_time in range(zeta.shape[0]):
    for i_lev in range(siglay.shape[0]):
        depth[i_time,i_lev,:] = -siglay[i_lev,:]*depth_all[i_time,:] # :是网格点

# 计算网格深度，现在为0时刻
print('计算cell深度')
depth_current = np.zeros((zeta.shape[0], siglay.shape[0], len(latc)), dtype=float) # (时间,层,网格)
for i_lev in range(siglay.shape[0]):
    print(i_lev)
    for i_cell_subfix in range(len(latc)):
        depth_current[0,i_lev,i_cell_subfix] = np.average([
            depth[11,i_lev,nv[0,i_cell_subfix]-1], # 格点号要-1
            depth[11,i_lev,nv[1,i_cell_subfix]-1],
            depth[11,i_lev,nv[2,i_cell_subfix]-1]
        ])
        depth_current[:,i_lev,i_cell_subfix] = depth_current[0,i_lev,i_cell_subfix]

# 筛选区域内的网格
lon_max = max(lon_adcp)+0.5
lon_min = min(lon_adcp)-0.5
lat_max = max(lat_adcp)+0.5
lat_min = min(lat_adcp)-0.5
latc_used = []      # latc(441027)->(31776)
lonc_used = []      # lonc(441027)->(31776)
u_used = []         # u(24,68,441027)->(31776,68)时间，层，格点
v_used = []         # v(24,68,441027)->(31776,68)
depth_used = []     # depth(24,68,226031)->(31776,68)
for i in range(len(latc)): # 网格相关参数
    if lonc[i] < lon_max and lonc[i] > lon_min and latc[i] > lat_min and latc[i] < lat_max:
        latc_used.append(latc[i])
        lonc_used.append(lonc[i])
        u_used.append(u[6,:,i])
        v_used.append(v[6,:,i])
        depth_used.append(depth_current[6,:,i])
    else:
        continue
latc_used = np.array(latc_used)     # (31776)
lonc_used = np.array(lonc_used)     # (31776)
u_used = np.array(u_used)           # (31776,68)
v_used = np.array(v_used)           # (31776,68)
depth_used = np.array(depth_used)   # (31776,68)

# 找到最近的网格
distance = []
for i_point in range(len(lat_adcp)): # 观测点
    distance.append([])
    print('正在处理格点：'+str(i_point+1)+'/151')
    for i_cell in range(len(latc_used)): # 网格
        distance[i_point].append(gd((lat_adcp[i_point], lon_adcp[i_point]), (latc_used[i_cell], lonc_used[i_cell])).km)
distance = np.array(distance)
# 计算离观测点最近的网格下标（筛选后的下标），从depth_used中获取
min_index = []
for i_point in range(len(lat_adcp)): # 观测点
    min_index.append(np.argmin(distance[i_point]))
min_index = np.array(min_index)

# 插值到指定深度
u_fvcom = []
v_fvcom = []
for i_point in range(len(min_index)):  # 插值:(目标点/一维数组,x,y) 获取对应深度的速度
    u_fvcom.append(np.interp(depth_adcp,depth_used[min_index[i_point],:],u_used[min_index[i_point],:],left=np.nan,right=np.nan)) # depth_used (31776,68)
    v_fvcom.append(np.interp(depth_adcp,depth_used[min_index[i_point],:],v_used[min_index[i_point],:],left=np.nan,right=np.nan))
u_fvcom = np.array(u_fvcom)  # (151,67) 格点，深度
v_fvcom = np.array(v_fvcom)  # (151,67)
u_fvcom *= 100
v_fvcom *= 100

# 绘图
# 创建colorbar
# 定义颜色
color_start = np.array([0.8941, 0.7529, 0.4235])
color_mid = np.array([1, 1, 1])
color_end = np.array([0, 0.1294, 0.7098])
# 定义两段的数量
n1 = 8
n2 = 28
total_n = n1 + n2
pos_mid = n1 / total_n
# 构建segmentdata
cdict = {
    'red': [
        (0, color_start[0], color_start[0]),
        (pos_mid, color_mid[0], color_mid[0]),
        (1, color_end[0], color_end[0])
    ],
    'green': [
        (0, color_start[1], color_start[1]),
        (pos_mid, color_mid[1], color_mid[1]),
        (1, color_end[1], color_end[1])
    ],
    'blue': [
        (0, color_start[2], color_start[2]),
        (pos_mid, color_mid[2], color_mid[2]),
        (1, color_end[2], color_end[2])
    ]
}
# 创建colormap，总离散级别数设为total_n（也可以设大一些，比如256，但位置比例不变）
my_cmap = mcolors.LinearSegmentedColormap('my_cmap', cdict, N=total_n)

# 读取水深数据
gebco_file = r'D:\Documents\Data\gebco_2024\GEBCO_2024.nc'
depth_data = netCDF4.Dataset(gebco_file, 'r')
depth_lon = depth_data.variables['lon'][:]
depth_lat = depth_data.variables['lat'][:]
depth_elevation = depth_data.variables['elevation'][:]
lon_used = depth_lon[np.argwhere(depth_lon > lon_min)[0][0]:np.argwhere(depth_lon < lon_max)[-1][0]]
lat_used = depth_lat[np.argwhere(depth_lat > lat_min)[0][0]:np.argwhere(depth_lat < lat_max)[-1][0]]
depth_used = depth_elevation[np.argwhere(depth_lat > lat_min)[0][0]:np.argwhere(depth_lat < lat_max)[-1][0],
                             np.argwhere(depth_lon > lon_min)[0][0]:np.argwhere(depth_lon < lon_max)[-1][0]]

for i_lev in range(len(depth_adcp)): # 深度层
    str_depth_adcp = str(depth_adcp[i_lev]) # 准备工作：准备字符串的深度数据
    fig = plt.figure(figsize=(12, 6))
    ax = fig.add_axes([0.1, 0.1, 0.8, 0.7], projection=ccrs.PlateCarree())
    ax.set_extent([lon_min, lon_max, lat_min, lat_max], crs=ccrs.PlateCarree())
    # 添加地图要素
    im = ax.pcolormesh(lon_used, lat_used, -depth_used,
                       cmap=my_cmap,
                       vmin=0, vmax=2000,
                       shading='auto',  # 自动处理网格点
                       transform=ccrs.PlateCarree())  # 确保正确的坐标变换

    # 创建与地图宽度对齐的底部 colorbar
    # 计算地图轴的位置和尺寸
    ax_pos = ax.get_position()
    cax_width = 0.05  # 与地图宽度相同
    cax_height = 0.62  # colorbar 高度
    cax_left = 0.92  # 与地图左侧对齐
    cax_bottom = 0.14  # 在地图下方，留出5%的间距

    # 创建 colorbar 轴
    cax = fig.add_axes([cax_left, cax_bottom, cax_width, cax_height])
    cbar = fig.colorbar(im, cax=cax, orientation='vertical')  # 垂直方向

    # 确保颜色范围设置正确
    im.set_clim(0, 2000)

    # 设置刻度
    cbar.ax.yaxis.set_major_locator(MultipleLocator(500))  # 水平 colorbar 使用 xaxis
    cbar.ax.yaxis.set_minor_locator(MultipleLocator(100))  # 水平 colorbar 使用 xaxis
    cbar.ax.tick_params(axis='y', which='minor', labelbottom=False)  # 隐藏副刻度标签

    # 添加单位（m）到右侧
    cbar.ax.set_title("m", loc="right", fontsize=10, pad=5)  # 调整位置和间距

    # 创建陆地掩码
    land_mask = (depth_used >= 0.)  # 深度≥0表示陆地

    # 创建陆地数据（1表示陆地，NaN表示海洋）
    land_data = np.where(land_mask, 1, np.nan)

    # 使用绿色填充陆地
    land_cmap = mcolors.ListedColormap(['green'])
    ax.pcolormesh(lon_used, lat_used, land_data,
                  cmap=land_cmap,
                  shading='auto',
                  transform=ccrs.PlateCarree(),
                  alpha=0.7)  # 设置透明度使地图更美观
    gl = ax.gridlines(draw_labels=True, linewidth=0.5, color='gray', alpha=0.5)
    gl.bottom_labels = False
    gl.right_labels = False
    ax.plot(lon_adcp,lat_adcp)
    # ax.quiver(lon_adcp, lat_adcp,u_adcp[i_lev,:],v_adcp[i_lev,:],scale=1000,width=0.001,color='blue', label='走航ADCP实测结果')
    # ax.quiver(lon_adcp, lat_adcp,u_fvcom[:,i_lev],v_fvcom[:,i_lev],scale=1000,width=0.001,color='red', label='FVCOM模型结果')
    # ax.legend()
    q1 = ax.quiver(lon_adcp, lat_adcp, u_adcp[i_lev, :], v_adcp[i_lev, :], scale=1000, width=0.001, color='blue')
    q2 = ax.quiver(lon_adcp, lat_adcp, u_fvcom[:, i_lev], v_fvcom[:, i_lev], scale=1000, width=0.001, color='red')

    # 添加自定义图例（箭头）
    ax.quiverkey(q1, 0.85, 0.92, 20, '走航ADCP实测', coordinates='axes', labelpos='E', fontproperties={'size': 10})
    ax.quiverkey(q2, 0.85, 0.87, 20, 'FVCOM模型结果', coordinates='axes', labelpos='E', fontproperties={'size': 10})
    ax.set_title('船载ADCP与FVCOM模型'+str_depth_adcp+'m深度流速结果对比')
    plt.savefig('output/compare_adcp/船载ADCP与FVCOM模型'+str_depth_adcp+'m深度流速结果对比.png', dpi=600, bbox_inches='tight')
    plt.close()