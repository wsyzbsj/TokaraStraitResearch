# 运行前先运行爬虫获取标准水位数据
import openpyxl
import glob, os, pathlib
import numpy as np
import pandas as pd

workbook = openpyxl.load_workbook('验潮站标准水位.xlsx')
# 读取站号
storage_paths = sorted(glob.glob('/home/yzbsj/Data/海洋数据/JODC/验潮站数据/*'))
for i_storage_path,storage_path in enumerate(storage_paths):
    if os.path.isdir(storage_path):
        sta_id = pathlib.Path(storage_path).stem
        print(sta_id)
        # 读取基准水位
        worksheet = workbook[sta_id]
        # 获取站号与站点编号
        if 'Code' in worksheet.cell(row=1,column=1).value:
            sta_code = worksheet.cell(row=1,column=2).value
        if 'Name' in worksheet.cell(row=2,column=1).value:
            sta_name = worksheet.cell(row=2,column=2).value
        # 获取基准高度行数
        begin_row = np.nan
        for i_row in range(1,worksheet.max_row+1):
            if worksheet.cell(row=i_row, column=1).value == 'Date':
                begin_row = i_row
                break
        dataframe_fix = pd.DataFrame()
        dataframe_fix['Time'] = ''
        dataframe_fix['latitude'] = ''
        dataframe_fix['longitude'] = ''
        # 读取
        for i_row in range(begin_row,worksheet.max_row+1):
            year_temp = worksheet.cell(row=i_row, column=1).value
            if year_temp=='':                       # 没有年份/日期
                temp = worksheet.cell(row=i_row, column=2).value
                if 'Zero of Tide Height' in temp:   # 基准水位行
                    temps = temp.split(' ')
                    if 'bench' in temp and 'mark' in temp:                # 局部基准点
                        for i_temp,temp_str in enumerate(temps):
                            if temp_str == 'Below':
                                while i_temp >0:
                                    if 'cm' in temps[i_temp-1]:
                                        dataframe_fix.loc[i_row,'bench mark'] = float(temps[i_temp-1].replace('cm',''))
                                        break
                                    else:
                                        i_temp -= 1
                    elif 'Standard Mark' in temp:           # 全局基准点
                        dataframe_fix.loc[i_row,'standard mark'] =
                    elif 'Tokyo Peil' in temp or 'T.P.' in temp:    # 东京
                        dataframe_fix.loc[i_row,'Tokyo Peil'] =
                    else:
                        continue
                elif 'Station Position' in temp:    # 站位信息
                    datas = temp.split(' ')
                    for i_data,data in enumerate(datas):
                        if data == 'N':
                            latitudes = datas[i_data-1].split('-')
                            latitude = float(latitudes[0])+float(latitudes[1])/60+float(latitudes[2])/3600
                        elif data == 'E':
                            longitudes = datas[i_data-1]
                            longitude = float(longitudes[0])+float(longitudes[1])/60+float(longitudes[2])/3600
                else:                               # 其他
                    continue
            else:                                   # 有年份/日期

                i_water = 0
    # 读取水位信息
    files = glob.glob(os.path.join(storage_path,'*.txt'))
    dataframe = pd.DataFrame()
    dataframe['站号'] = ''
    dataframe['时间'] = ''
    dataframe['水位'] = ''
    i_row = 0
    for file in files:
        with open(file,'r') as fin:
            # 获取总长度
            text = fin.read()
            char_count = len(text)
            fin.seek(0)
            i_count = 0
            i_hour = 0
            date = ''
            data = ''
            while i_count < char_count:
                temp = fin.read(1)
                if temp == ',':
                    if i_hour == 0: # 站号
                        sta_id = data
                    elif i_hour ==1:# 日期
                        date = data
                    else:   # 0、1为站号和日期
                        dataframe.loc[i_row,'站号'] = sta_id
                        dataframe.loc[i_row,'时间'] = date+' '+f"{i_hour-2:02d}"+":00:00"
                        dataframe.loc[i_row,'水位'] = data
                        i_row +=1
                    data = ''
                    i_hour += 1     # 小时+1(-2为最终真实时间)
                    i_count +=1
                elif temp == '\n':
                    dataframe.loc[i_row,'站号'] = sta_id
                    dataframe.loc[i_row,'时间'] = date+' '+f"{i_hour-2:02d}"+":00:00"
                    dataframe.loc[i_row,'水位'] = data
                    i_row +=1
                    data = ''
                    date = ''
                    i_hour = 0      # 换行表示时间为0时
                    i_count +=1
                else:
                    data += temp
                    i_count +=1
        print(1)