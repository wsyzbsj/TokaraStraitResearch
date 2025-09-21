import glob
import pathlib
import os
from zipfile import ZipFile
import logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
file_read = logging.StreamHandler()
file_read.setLevel(logging.INFO)
logger.addHandler(file_read)
import pathlib
import coare36vnWarm_et
import pandas as pd
import openpyxl

# 打开写入Excel
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.cell(row=1,column=1).value='时间'
worksheet.cell(row=1,column=2).value='航次号'
worksheet.cell(row=1,column=3).value='站点号'
worksheet.cell(row=1,column=4).value='纬度'
worksheet.cell(row=1,column=5).value='经度'
worksheet.cell(row=1,column=6).value='采样深度(m)'
worksheet.cell(row=1,column=7).value='温度(OBS)'
worksheet.cell(row=1,column=8).value='盐度(OBS)'
worksheet.cell(row=1,column=9).value='DEPTH(STD)'
worksheet.cell(row=1,column=10).value='TEMPERATURE(STD)'
worksheet.cell(row=1,column=11).value='SAL(STD)'
i_excel_write=2

# 读取盐度温度数据
logger.info('Hydrographic data process begin.')
e_files = sorted(glob.glob('/home/yzbsj/Data/气象数据/Oceanographic and Marine Meteorological Observations by Research Vessels/*/*/*/海洋数据/*.E'))
for e_file in e_files:
    with open(e_file) as fin:
        filename = pathlib.Path(e_file).name
        # 获取文件尾位置
        fin.seek(0,2)
        fin_eof = fin.tell()
        fin.seek(0,0)
        logger.info('File ' + e_file + ' processing begin.')
        i_line = 1
        # Line 1 —— 航次信息
        data_format = fin.read(5)
        cruise_no = fin.read(5).strip()
        period = fin.read(10)
        if data_format.startswith('E2.'):
            case =True
            area = fin.read(98)
        elif data_format.startswith('E3.'):
            case = False
            area = fin.read(113)
        else:
            logger.error('File '+e_file+' line '+str(i_line)+" data format error")
            exit(-1)
        station_count = fin.read(5)
        ship_code = fin.read(2)
        temp_0 = fin.read(2) # @\n
        latitude   = [] # 观测纬度
        longitude  = [] # 观测经度
        obs_time   = [] # 观测时间(UTC+9)
        water_depth= [] # 水深
        water_color= [] # 水色
        water_trans= [] # 透明度
        ssf_no     = [] # 次表层水温数据对应站号
        acm_no     = [] # 次表层洋流数据对应站号
        sub_sta_no = [] # 子站号
        logger.debug('File '+e_file + ' line ' + str(i_line) + ' process complete.')
        i_line += 1
        # Line 2 —— 站位信息
        while fin.tell() < fin_eof:
            station_no = fin.read(8)
            latitude.append(fin.read(8))
            longitude.append(fin.read(9))
            obs_time.append(fin.read(22))
            if case:                                # E2.x
                water_depth.append(fin.read(6))
                water_color.append(fin.read(3))
                water_trans.append(fin.read(45))
            else:                                   # E3.x
                water_depth.append(fin.read(68))
            ssf_no.append(fin.read(7))
            acm_no.append(fin.read(7))
            if case:
                sub_sta_no.append(fin.read(6))
            elif pathlib.Path(e_file).name=='RF2005.E' and i_line == 569: # RF2005.E第569行存在格式问题，多一个字符，单独处理
                sub_sta_no.append(fin.read(8))
            else:
                sub_sta_no.append(fin.read(7))
            cruise_no = fin.read(4)
            temp_1 = fin.read(2)                    # =\n
            if temp_1 != '=\n':
                logger.error('File '+e_file + ' Line ' + str(i_line) + ' endding character not correct.')
                exit(-1)
            logger.debug('File '+e_file + ' line ' + str(i_line) + ' process complete.')
            i_line += 1
            # Line 3 —— Station Remarks
            station_no = fin.read(8)
            remarks = fin.read(82)                  # remarks
            if case:
                param_inf = fin.read(35)            # param_inf  可选参数定义
            else:
                param_inf = fin.read(50)
            param_def=param_inf.strip()
            if len(param_def) == 0:                 # 未定义可选参数
                pass
            else:
                params = param_def.split(',')
                params_name = []
                params_column_begin = []
                params_column_end = []
                params_data = []
                for param in params:
                    temp = param.strip().split(' ')
                    params_name.append(temp[0].strip())
                    params_column_begin.append(temp[1].split('-')[0].strip())
                    params_column_end.append(temp[1].split('-')[1].strip())
                    params_data.append([])
            temp_2 = fin.read(2)                    # =\n
            if temp_2 != '=\n':
                logger.error('File '+e_file + ' Line ' + str(i_line) + ' endding character not correct.')
                exit(-1)
            logger.debug('File '+e_file + ' file header read complete.')
            i_line += 1
            # 数据行
            bool_data_end = True
            while bool_data_end:
                station_no_current = fin.read(8)    # STATION NO 站号
                time = fin.read(8)                  # TIME       采样时间(日本时间, UTC+9)
                depth_obs = fin.read(5)             # DEPTH(OBS) 采样深度(m)
                # E-2.x格式
                if case:
                    temp_obs = fin.read(6)          # TEMP(OBS)  温度——CTD温度, 1990年国际温标 (ITS-90)
                    sal_obs = fin.read(7)           # SAL(OBS)   实用盐度——CTD盐度, 1978年实用盐度表(PSS-78)
                    DO = fin.read(4)                # DO         溶解氧浓度(µmol/L)——Winkler 方法测定的溶解氧浓度
                    p04_p = fin.read(5)             # P04-P      磷酸根的磷(µmol/L)——(Strickland and Parsons, 1965)
                    t_p = fin.read(5)               # T-P        总磷含量(µmol/L)——过硫酸钾(K2S2O8)分解法测定
                    no3_n = fin.read(5)             # NO3-N      硝酸根的氮(µmol/L)——(Wood, Armstrong and Richard, 1967)
                    no2_n = fin.read(5)             # NO2-N      亚硝酸根的氮(µmol/L)——(Strickland and Parsons, 1965)
                    nh3_n = fin.read(5)             # NH3-N      铵的氮(µmol/L)
                    ph = fin.read(5)                # PH         PH值
                    chl = fin.read(7)               # CHL        叶绿素a含量(µg/L)——荧光法测定
                    pha = fin.read(7)               # PHA        浅色素含量(µg/L)——荧光法测定
                    add_para = fin.read(11)         # ADD PARAM
                    temp_add_data = []
                    for i_para_name,param_name in enumerate(params_name):
                        temp_add_data.append(add_para[int(params_column_begin[i_para_name])-83:int(params_column_end[i_para_name])-82])
                    params_data.append(temp_add_data)
                    depth_std = fin.read(5)         # DEPTH(STD) Standard depths in meters.
                    temp_std = fin.read(6)          # TEMP(STD)  CTD temperature in ITS-90.
                    sal_std = fin.read(11)          # SAL(STD)   CTD salinity in PSS-78.
                    d_st = fin.read(5)              # D-ST       Thermosteric anomaly in 10−8 m3/kg.
                # E-3.x格式
                else:
                    temp_obs = fin.read(7)          # TEMP(OBS)
                    sal_obs = fin.read(7)           # SAL(OBS)   实用盐度——(1978年实用盐度表,PSS-78)中的CTD盐度
                    DO = fin.read(6)                # DO         溶解氧浓度(µmol/kg)——Winkler 方法测定的溶解氧浓度
                    po4_p = fin.read(6)             # PO4-P      磷酸根的磷(µmol/kg) (Strickland and Parsons, 1965)
                    no3_n = fin.read(6)             # NO3-N      硝酸根的氮(µmol/kg)——(Wood, Armstrong and Richard, 1967)
                    no2_n = fin.read(5)             # NO2-N      亚硝酸根的氮(µmol/kg)——(Strickland and Parsons, 1965)
                    slica = fin.read(6)             # SLICA      硅酸盐硅的含量(µmol/kg)——(Grasshoff et al, 1983)
                    ph = fin.read(6)                # PH         PH值
                    chl = fin.read(7)               # CHL        叶绿素a含量(µg/L)——荧光法测定
                    pha = fin.read(7)               # PHA        浅色素含量(µg/L)——荧光法测定
                    add_para = fin.read(21)         # ADD PARAM
                    temp_add_data = []
                    for i_para_name,param_name in enumerate(params_name):
                        temp_add_data.append(add_para[int(params_column_begin[i_para_name])-85:int(params_column_end[i_para_name])-84])
                    params_data.append(temp_add_data)
                    depth_std = fin.read(5)         # DEPTH(STD) Standard depths in meters.
                    temp_std = fin.read(7)          # TEMP(STD)  CTD temperature in ITS-90.
                    sal_std = fin.read(7)           # SAL(STD)   CTD salinity in PSS-78.
                    do = fin.read(6)                # DO         CTD oxygen in micromoles per kilogram.
                    d_st = fin.read(5)              # D-ST       Thermosteric anomaly in 10-8 m3/kg.
                delta_d = fin.read(5)               # DELTA-D    重力异常10m2/s2
                temp_3 = fin.read(2)                # REC_IND    @\n或=\n
                worksheet.cell(row=i_excel_write,column=1).value = obs_time[-1]             # 时间
                worksheet.cell(row=i_excel_write,column=2).value = filename.split('.')[0]   # 航次号
                worksheet.cell(row=i_excel_write,column=3).value = station_no               # 站点号
                worksheet.cell(row=i_excel_write,column=4).value = latitude[-1]             # 纬度
                worksheet.cell(row=i_excel_write,column=5).value = longitude[-1]            # 经度
                worksheet.cell(row=i_excel_write,column=6).value = depth_obs                # 深度(OBS)
                worksheet.cell(row=i_excel_write,column=7).value = temp_obs                 # 温度(OBS)
                worksheet.cell(row=i_excel_write,column=8).value = sal_obs                  # 盐度(OBS)
                worksheet.cell(row=i_excel_write,column=9).value = depth_std                # 深度(STD)
                worksheet.cell(row=i_excel_write,column=10).value = temp_std                # 温度(STD)
                worksheet.cell(row=i_excel_write,column=11).value = sal_std                 # 盐度(STD)
                i_excel_write += 1
                if temp_3 == '=\n':
                    bool_data_end = True
                elif temp_3 == '@\n':
                    bool_data_end = False
                else:
                    logger.error('File '+e_file + ' Line ' + str(i_line) +' line end character not correct.')
                    exit(-1)
                logger.debug('File '+e_file + ' Line ' + str(i_line) + ' process complete.')
                i_line += 1
    logger.info('File '+e_file+' processing complete.')
logger.info('Hydrographic data process complete.')
workbook.save('水文数据.xlsx')
logger.info('Excel file output complete.')

# 读取CTD
logger.info('CTD data process begin.')
ctd_files = sorted(glob.glob('/home/yzbsj/Data/气象数据/Oceanographic and Marine Meteorological Observations by Research Vessels/*/*/*/海洋数据/*.CTD*'))
for ctd_file in ctd_files:
    filetype = pathlib.Path(ctd_file).suffix
    filename = pathlib.Path(ctd_file).name
    logger.info('File ' + filename + ' processing begin.')
    with ZipFile(ctd_file, 'r') as fin:
        fin.extractall('ctd_cache')
        os.mkdir('ctd_cache')
        logger.info('File ' + filename + ' processing end.')

        logger.info('File ' + filename + ' processing end.')
    logger.info('File ' + filename + ' processing end.')
logger.info('CTD data process end.')

# 读取XCTD数据
logger.info('XCTD data process begin.')
xctd_files = sorted(glob.glob('/home/yzbsj/Data/气象数据/Oceanographic and Marine Meteorological Observations by Research Vessels/*/*/*/海洋数据/*.XCT*'))
if xctd_files:
    pass
else:
    logger.info('No XCTD data found.')
logger.info('XCTD data process end.')